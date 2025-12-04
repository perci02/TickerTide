from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from webdriver_manager.chrome import ChromeDriverManager

from datetime import datetime
import pandas as pd
import time
import os

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment


# -------------------------------------------------
# Browser setup
# -------------------------------------------------
def get_driver(headless=False):
    """Create and return a Chrome WebDriver instance."""
    options = Options()
    if headless:
        options.add_argument("--headless=new")

    options.add_argument("--start-maximized")
    options.add_argument("--disable-gpu")
    options.add_argument("--no-sandbox")
    options.add_argument("--disable-dev-shm-usage")

    service = Service(ChromeDriverManager().install())
    driver = webdriver.Chrome(service=service, options=options)
    return driver


# -------------------------------------------------
# Scraping logic
# -------------------------------------------------
def scrape_top_60_coins(driver):
    """Scrape top 60 cryptocurrencies from CoinMarketCap."""
    url = "https://coinmarketcap.com/"
    driver.get(url)

    # Wait for page and JS to fully load
    time.sleep(8)

    rows = driver.find_elements(By.CSS_SELECTOR, "table tbody tr")
    rows = rows[:60]

    data = []
    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    for row in rows:
        cols = row.find_elements(By.TAG_NAME, "td")
        if len(cols) < 8:
            continue

        try:
            name = cols[1].text.split("\n")[0]
            price = cols[2].text
            change_24h = cols[4].text
            market_cap = cols[6].text

            data.append({
                "timestamp": timestamp,
                "name": name,
                "price": price,
                "change_24h": change_24h,
                "market_cap": market_cap
            })
        except Exception as e:
            print("Error parsing row:", e)
            continue

    return data


# CSV saving (raw data)

def save_to_csv(data, filename="crypto_prices.csv"):
    """Save scraped data into CSV (append mode)."""
    if not data:
        print("No data to save to CSV.")
        return

    df = pd.DataFrame(data)
    file_exists = os.path.isfile(filename)
    df.to_csv(filename, mode="a", index=False, header=not file_exists)
    print(f"Saved {len(data)} records to {filename}")


# -------------------------------------------------
# Excel saving (styled, append)
# -------------------------------------------------
def save_to_excel_styled(data, filename="crypto_prices.xlsx"):
    """Append data to a styled Excel file (create if not exists)."""
    if not data:
        print("No data to save to Excel.")
        return

    headers = ["Timestamp", "Name", "Price", "24h Change", "Market Cap"]

    # Styles
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )
    center_align = Alignment(horizontal="center")

    if os.path.exists(filename):
        # File exists -> open and append
        wb = load_workbook(filename)
        ws = wb.active
        print(f"Appending to existing Excel file: {filename}")

        start_row = ws.max_row + 1

        # Append new rows
        for item in data:
            row = [
                item["timestamp"],
                item["name"],
                item["price"],
                item["change_24h"],
                item["market_cap"]
            ]
            ws.append(row)

        # Apply border + alignment for only new rows
        for row in ws.iter_rows(
            min_row=start_row,
            max_row=ws.max_row,
            min_col=1,
            max_col=len(headers)
        ):
            for cell in row:
                cell.border = thin_border
                cell.alignment = center_align

    else:
        # File doesn't exist -> create new with header styling
        wb = Workbook()
        ws = wb.active
        ws.title = "Crypto Data"
        print(f"Creating new Excel file: {filename}")

        # Header row
        ws.append(headers)

        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(
            start_color="4F81BD",
            end_color="4F81BD",
            fill_type="solid"
        )

        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = thin_border

        # Data rows
        for item in data:
            row = [
                item["timestamp"],
                item["name"],
                item["price"],
                item["change_24h"],
                item["market_cap"]
            ]
            ws.append(row)

        # Style all data rows
        for row in ws.iter_rows(
            min_row=2,
            max_row=ws.max_row,
            min_col=1,
            max_col=len(headers)
        ):
            for cell in row:
                cell.border = thin_border
                cell.alignment = center_align

    # Auto-adjust column widths
    for col in ws.columns:
        max_len = 0
        column_letter = col[0].column_letter
        for cell in col:
            val = "" if cell.value is None else str(cell.value)
            if len(val) > max_len:
                max_len = len(val)
        ws.column_dimensions[column_letter].width = max_len + 2

    wb.save(filename)
    print(f"Styled Excel file updated: {filename}")


# -------------------------------------------------
# Helper functions for filters (optional)
# -------------------------------------------------
def parse_price(price_str):
    """'$42,000.12' -> 42000.12"""
    clean = price_str.replace("$", "").replace(",", "").strip()
    try:
        return float(clean)
    except:
        return None


def parse_percent(percent_str):
    """'-2.45%' -> -2.45"""
    clean = percent_str.replace("%", "").replace(",", "").strip()
    try:
        return float(clean)
    except:
        return None


def filter_by_min_price(coins, min_price):
    """Filter coins with price >= min_price."""
    result = []
    for coin in coins:
        p = parse_price(coin["price"])
        if p is not None and p >= min_price:
            result.append(coin)
    return result


def filter_top_gainers_24h(coins, min_change):
    """Filter coins with 24h change >= min_change."""
    result = []
    for coin in coins:
        ch = parse_percent(coin["change_24h"])
        if ch is not None and ch >= min_change:
            result.append(coin)
    return result


# -------------------------------------------------
# Main
# -------------------------------------------------
def main():
    driver = get_driver(headless=False)   # change to True for headless mode

    try:
        coins = scrape_top_60_coins(driver)

        print("Top 60 coins scraped:")
        for coin in coins:
            print(coin)

        # Optional: examples of filters (for terminal output only)
        high_price = filter_by_min_price(coins, 10000)
        print("\nCoins with price >= $10,000:")
        for coin in high_price:
            print(coin["name"], coin["price"])

        big_gainers = filter_top_gainers_24h(coins, 5)
        print("\nCoins with 24h gain >= 5%:")
        for coin in big_gainers:
            print(coin["name"], coin["change_24h"])

        # Save data
        save_to_csv(coins)              # raw, historical
        save_to_excel_styled(coins)     # styled, historical

    finally:
        driver.quit()


if __name__ == "__main__":
    main()
